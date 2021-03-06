import openpyxl                                         # To work with Excel Files
from selenium import webdriver

# from inputData import getSheet, getRegNoData            # To handle all inputs
from result import getAllSemesterResult                 # To fetch results


# Main Module -->

def puResult(book, sheet, filePath, driver, col, rowFrom, rowTo, semFrom, semTo, screenCap):
  driver.get('http://exam.pondiuni.edu.in/results/')
  driver.maximize_window()
  newCols = []
  for currentRow in range(rowFrom, rowTo+1):
    currentRegNo = sheet.cell(row = currentRow, column = col).value
    allSemResult, driver = getAllSemesterResult(driver, currentRegNo, semFrom, semTo, screenCap)
    colCount = sheet.max_column
    for subject in allSemResult:
      for row in sheet.iter_rows():
        for cell in row:
          if cell.value == subject:
            sheet.cell(row = currentRow, column = cell.column).value = allSemResult[subject]
            allSemResult[subject] = None
    for subject in allSemResult:
      if allSemResult[subject] != None:
        colCount = colCount + 1
        sheet.cell(row = 1, column = colCount).value = subject
        sheet.cell(row = currentRow, column = colCount).value = allSemResult[subject]
        newCols.append('The Subject "' + subject + '" not present in the table. So, it is added in the last')
        
  book.save(filePath)
  driver.quit()
  return newCols